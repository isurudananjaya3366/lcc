"""Views for ReorderSuggestion with conversion and reporting actions."""

import csv
import logging
from datetime import timedelta
from io import BytesIO

from django.db import transaction
from django.db.models import Avg, Count, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django_filters import rest_framework as django_filters
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.inventory.alerts.models import ReorderSuggestion
from apps.inventory.alerts.serializers import (
    ReorderSuggestionListSerializer,
    ReorderSuggestionSerializer,
)

logger = logging.getLogger(__name__)


# ── Filters ─────────────────────────────────────────────────────────


class ReorderSuggestionFilter(django_filters.FilterSet):
    """Filter set for ReorderSuggestion."""

    urgency = django_filters.CharFilter(field_name="urgency")
    status = django_filters.CharFilter(field_name="status")
    warehouse = django_filters.UUIDFilter(field_name="warehouse_id")
    supplier = django_filters.UUIDFilter(field_name="suggested_supplier_id")
    product = django_filters.UUIDFilter(field_name="product_id")
    min_cost = django_filters.NumberFilter(field_name="estimated_cost", lookup_expr="gte")
    max_cost = django_filters.NumberFilter(field_name="estimated_cost", lookup_expr="lte")
    is_expired = django_filters.BooleanFilter(method="filter_expired")

    class Meta:
        model = ReorderSuggestion
        fields = ["urgency", "status", "warehouse", "supplier", "product"]

    def filter_expired(self, queryset, name, value):
        cutoff = timezone.now() - timedelta(days=30)
        if value:
            return queryset.filter(status="pending", created_at__lt=cutoff)
        return queryset.exclude(status="pending", created_at__lt=cutoff)


# ── ViewSet ─────────────────────────────────────────────────────────


class ReorderSuggestionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet with conversion, dismiss, report, and bulk actions.

    Endpoints:
        GET  /reorder/                        - List
        GET  /reorder/{id}/                   - Retrieve
        POST /reorder/{id}/convert_to_po/     - Convert to PO
        POST /reorder/{id}/dismiss/           - Dismiss
        POST /reorder/bulk_convert/           - Bulk convert
        GET  /reorder/summary/                - Summary stats
        GET  /reorder/report/                 - Report (JSON/CSV/Excel)
        POST /reorder/email_report/           - Queue email report
    """

    queryset = ReorderSuggestion.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ReorderSuggestionFilter
    search_fields = ["product__name", "product__sku", "notes"]
    ordering_fields = [
        "created_at",
        "days_until_stockout",
        "estimated_cost",
        "urgency",
    ]
    ordering = ["days_until_stockout"]

    def get_serializer_class(self):
        if self.action == "list":
            return ReorderSuggestionListSerializer
        return ReorderSuggestionSerializer

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("product", "warehouse", "suggested_supplier")
        )

    # ── Lifecycle actions ───────────────────────────────────────

    @action(detail=True, methods=["post"])
    def convert_to_po(self, request, pk=None):
        """
        Convert suggestion to Purchase Order (if purchasing module exists).

        Body: {"supplier_id": ..., "notes": ..., "delivery_date": ...}
        """
        suggestion = self.get_object()
        can, reason = suggestion.can_convert()
        if not can:
            return Response({"error": reason}, status=status.HTTP_400_BAD_REQUEST)

        supplier_id = request.data.get("supplier_id") or suggestion.suggested_supplier_id
        if not supplier_id:
            return Response(
                {"error": "Supplier is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from apps.purchasing.services.po_converter import POConverter

            with transaction.atomic():
                po = POConverter.convert_suggestion_to_po(
                    suggestion=suggestion,
                    supplier_id=supplier_id,
                    notes=request.data.get("notes"),
                    delivery_date=request.data.get("delivery_date"),
                )
                suggestion.mark_converted(po_id=po.id, user=request.user)

            return Response(
                {
                    "message": "Suggestion converted to Purchase Order",
                    "suggestion": ReorderSuggestionSerializer(suggestion).data,
                }
            )
        except ImportError:
            # Purchasing module not available — mark converted with a placeholder UUID
            import uuid

            with transaction.atomic():
                suggestion.mark_converted(po_id=uuid.uuid4(), user=request.user)
            return Response(
                {
                    "message": "Suggestion marked as converted (purchasing module unavailable)",
                    "suggestion": ReorderSuggestionSerializer(suggestion).data,
                }
            )
        except Exception as exc:
            logger.exception("Error converting suggestion %s to PO", suggestion.id)
            return Response(
                {"error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["post"])
    def dismiss(self, request, pk=None):
        """
        Dismiss a pending suggestion.

        Body: {"reason": "Already ordered manually"}
        """
        suggestion = self.get_object()
        if suggestion.status != "pending":
            return Response(
                {"error": "Only pending suggestions can be dismissed"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        reason = request.data.get("reason", "")
        suggestion.mark_dismissed(reason=reason, user=request.user)
        return Response(ReorderSuggestionSerializer(suggestion).data)

    @action(detail=False, methods=["post"])
    def bulk_convert(self, request):
        """
        Convert multiple suggestions.

        Body: {"suggestion_ids": [...]}
        """
        suggestion_ids = request.data.get("suggestion_ids", [])
        if not suggestion_ids:
            return Response(
                {"error": "suggestion_ids list is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        suggestions = self.get_queryset().filter(id__in=suggestion_ids, status="pending")
        if not suggestions.exists():
            return Response(
                {"error": "No valid suggestions found"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from apps.purchasing.services.po_converter import POConverter

            with transaction.atomic():
                result = POConverter.bulk_convert_suggestions(suggestions=suggestions)
            return Response(
                {
                    "message": "Suggestions converted successfully",
                    "pos_created": result["pos_created"],
                    "suggestions_converted": result["suggestions_converted"],
                }
            )
        except ImportError:
            return Response(
                {"error": "Purchasing module not available"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # ── Reporting actions ───────────────────────────────────────

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Counts by urgency and status."""
        qs = self.filter_queryset(self.get_queryset())
        by_urgency = {}
        for item in (
            qs.filter(status="pending")
            .values("urgency")
            .annotate(count=Count("id"), total_cost=Sum("estimated_cost"))
        ):
            by_urgency[item["urgency"]] = {
                "count": item["count"],
                "total_cost": float(item["total_cost"] or 0),
            }

        return Response(
            {
                "total_pending": qs.filter(status="pending").count(),
                "total_converted": qs.filter(status="converted_to_po").count(),
                "total_dismissed": qs.filter(status="dismissed").count(),
                "by_urgency": by_urgency,
                "total_estimated_cost": float(
                    qs.filter(status="pending").aggregate(t=Sum("estimated_cost"))["t"] or 0
                ),
            }
        )

    @action(detail=False, methods=["get"])
    def report(self, request):
        """
        Comprehensive report with filtering and optional CSV / Excel export.

        Query params: export_format (json|csv|excel), urgency, status, warehouse,
        supplier, date_from, date_to, group_by, min_cost, max_cost.
        """
        export_format = request.query_params.get("export_format", "json")
        qs = self._apply_report_filters(request)

        summary = self._build_summary(qs)

        if export_format == "csv":
            return self._export_csv(qs, summary)
        if export_format == "excel":
            return self._export_excel(qs, summary)

        # JSON
        group_by = request.query_params.get("group_by")
        grouped = self._group_suggestions(qs, group_by) if group_by else {
            "ungrouped": ReorderSuggestionListSerializer(
                qs.order_by("days_until_stockout"), many=True
            ).data
        }
        return Response(
            {
                "summary": summary,
                "grouped_data": grouped,
                "generated_at": timezone.now(),
            }
        )

    @action(detail=False, methods=["post"])
    def email_report(self, request):
        """Queue email delivery of the reorder report."""
        email = request.data.get("email", request.user.email)
        if not email:
            return Response(
                {"error": "email is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        from apps.inventory.alerts.tasks import send_weekly_reorder_report

        send_weekly_reorder_report.delay()
        return Response({"message": "Report email queued", "email": email})

    @action(detail=False, methods=["get"])
    def calendar(self, request):
        """Return reorder calendar events for upcoming stockout predictions."""
        from apps.inventory.alerts.services.reports import ReorderCalendarService

        days_ahead = int(request.query_params.get("days_ahead", 30))
        days_ahead = min(max(days_ahead, 7), 90)  # Clamp 7-90

        service = ReorderCalendarService()
        events = service.get_calendar_events(days_ahead=days_ahead)

        return Response(
            {
                "days_ahead": days_ahead,
                "events": events,
                "total_events": len(events),
                "generated_at": timezone.now(),
            }
        )

    # ── Internal helpers ────────────────────────────────────────

    def _apply_report_filters(self, request):
        qs = self.get_queryset()
        params = request.query_params
        if params.get("status"):
            qs = qs.filter(status=params["status"])
        else:
            qs = qs.filter(status="pending")
        if params.get("urgency"):
            qs = qs.filter(urgency=params["urgency"])
        if params.get("warehouse"):
            qs = qs.filter(warehouse_id=params["warehouse"])
        if params.get("supplier"):
            qs = qs.filter(suggested_supplier_id=params["supplier"])
        if params.get("date_from"):
            d = parse_date(params["date_from"])
            if d:
                qs = qs.filter(created_at__date__gte=d)
        if params.get("date_to"):
            d = parse_date(params["date_to"])
            if d:
                qs = qs.filter(created_at__date__lte=d)
        if params.get("min_cost"):
            qs = qs.filter(estimated_cost__gte=params["min_cost"])
        if params.get("max_cost"):
            qs = qs.filter(estimated_cost__lte=params["max_cost"])
        return qs

    @staticmethod
    def _build_summary(qs):
        agg = qs.aggregate(
            total_cost=Sum("estimated_cost"),
            total_qty=Sum("suggested_qty"),
            avg_days=Avg("days_until_stockout"),
        )
        by_urgency = {}
        for item in qs.values("urgency").annotate(
            count=Count("id"), tc=Sum("estimated_cost")
        ):
            by_urgency[item["urgency"]] = {
                "count": item["count"],
                "total_cost": float(item["tc"] or 0),
            }
        return {
            "total_suggestions": qs.count(),
            "total_estimated_cost": float(agg["total_cost"] or 0),
            "total_quantity": float(agg["total_qty"] or 0),
            "avg_days_until_stockout": round(float(agg["avg_days"] or 0), 2),
            "by_urgency": by_urgency,
        }

    @staticmethod
    def _group_suggestions(qs, group_by):
        grouped = {}
        if group_by == "urgency":
            for level in ("critical", "high", "medium", "low"):
                sub = qs.filter(urgency=level)
                if sub.exists():
                    grouped[level] = {
                        "count": sub.count(),
                        "total_cost": float(
                            sub.aggregate(t=Sum("estimated_cost"))["t"] or 0
                        ),
                        "suggestions": ReorderSuggestionListSerializer(sub, many=True).data,
                    }
        elif group_by == "supplier":
            for item in qs.values("suggested_supplier_id", "suggested_supplier__name").distinct():
                name = item["suggested_supplier__name"] or "No Supplier"
                sub = qs.filter(suggested_supplier_id=item["suggested_supplier_id"])
                grouped[name] = {
                    "count": sub.count(),
                    "total_cost": float(
                        sub.aggregate(t=Sum("estimated_cost"))["t"] or 0
                    ),
                    "suggestions": ReorderSuggestionListSerializer(sub, many=True).data,
                }
        elif group_by == "category":
            for item in qs.values(
                "product__category_id", "product__category__name"
            ).distinct():
                name = item["product__category__name"] or "Uncategorized"
                sub = qs.filter(product__category_id=item["product__category_id"])
                grouped[name] = {
                    "count": sub.count(),
                    "total_cost": float(
                        sub.aggregate(t=Sum("estimated_cost"))["t"] or 0
                    ),
                    "suggestions": ReorderSuggestionListSerializer(sub, many=True).data,
                }
        elif group_by == "warehouse":
            for item in qs.values("warehouse_id", "warehouse__name").distinct():
                name = item["warehouse__name"] or "All Warehouses"
                sub = qs.filter(warehouse_id=item["warehouse_id"])
                grouped[name] = {
                    "count": sub.count(),
                    "total_cost": float(
                        sub.aggregate(t=Sum("estimated_cost"))["t"] or 0
                    ),
                    "suggestions": ReorderSuggestionListSerializer(sub, many=True).data,
                }
        return grouped

    @staticmethod
    def _export_csv(qs, summary):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="reorder_report_{timezone.now():%Y%m%d}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(
            [
                "Product SKU",
                "Product Name",
                "Warehouse",
                "Suggested Qty",
                "Current Stock",
                "Urgency",
                "Days Until Stockout",
                "Supplier",
                "Estimated Cost (LKR)",
                "Status",
                "Created At",
            ]
        )
        for s in qs.select_related("product", "warehouse", "suggested_supplier"):
            d = s.days_until_stockout
            writer.writerow(
                [
                    s.product.sku,
                    s.product.name,
                    s.warehouse.name if s.warehouse else "All",
                    float(s.suggested_qty),
                    float(s.current_stock),
                    s.urgency,
                    float(d) if d is not None else "",
                    s.suggested_supplier.name if s.suggested_supplier else "",
                    float(s.estimated_cost),
                    s.status,
                    s.created_at.strftime("%Y-%m-%d %H:%M"),
                ]
            )
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        writer.writerow(["Total Suggestions", summary["total_suggestions"]])
        writer.writerow(["Total Estimated Cost (LKR)", summary["total_estimated_cost"]])
        writer.writerow(["Total Quantity", summary["total_quantity"]])
        return response

    @staticmethod
    def _export_excel(qs, summary):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return Response(
                {"error": "openpyxl not installed for Excel export"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "Reorder Report Summary"
        ws["A1"].font = Font(size=14, bold=True)
        row = 3
        for label, val in [
            ("Total Suggestions", summary["total_suggestions"]),
            ("Total Estimated Cost (LKR)", summary["total_estimated_cost"]),
            ("Total Quantity", summary["total_quantity"]),
        ]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val)
            row += 1

        ws_d = wb.create_sheet("Details")
        headers = [
            "Product SKU",
            "Product Name",
            "Warehouse",
            "Suggested Qty",
            "Current Stock",
            "Urgency",
            "Days Until Stockout",
            "Supplier",
            "Estimated Cost",
            "Status",
            "Created At",
        ]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for c, h in enumerate(headers, 1):
            cell = ws_d.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r, s in enumerate(
            qs.select_related("product", "warehouse", "suggested_supplier"), 2
        ):
            d = s.days_until_stockout
            ws_d.cell(row=r, column=1, value=s.product.sku)
            ws_d.cell(row=r, column=2, value=s.product.name)
            ws_d.cell(row=r, column=3, value=s.warehouse.name if s.warehouse else "All")
            ws_d.cell(row=r, column=4, value=float(s.suggested_qty))
            ws_d.cell(row=r, column=5, value=float(s.current_stock))
            ws_d.cell(row=r, column=6, value=s.urgency)
            ws_d.cell(row=r, column=7, value=float(d) if d is not None else "")
            ws_d.cell(
                row=r,
                column=8,
                value=s.suggested_supplier.name if s.suggested_supplier else "",
            )
            ws_d.cell(row=r, column=9, value=float(s.estimated_cost))
            ws_d.cell(row=r, column=10, value=s.status)
            ws_d.cell(row=r, column=11, value=s.created_at.strftime("%Y-%m-%d %H:%M"))

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="reorder_report_{timezone.now():%Y%m%d}.xlsx"'
        )
        return resp
