"""
Excel report exporter using openpyxl.

Provides base utilities and report-specific export methods
for all financial reports.
"""

import logging
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List

from django.http import HttpResponse

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Styling ─────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF") if HAS_OPENPYXL else None
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid") if HAS_OPENPYXL else None
TITLE_FONT = Font(name="Arial", bold=True, size=16) if HAS_OPENPYXL else None
SUBTOTAL_FONT = Font(name="Arial", bold=True, size=10) if HAS_OPENPYXL else None
THIN_BORDER = Border(
    top=Side(style="thin"),
    bottom=Side(style="thin"),
    left=Side(style="thin"),
    right=Side(style="thin"),
) if HAS_OPENPYXL else None
DOUBLE_BORDER = Border(
    top=Side(style="double"),
    bottom=Side(style="double"),
) if HAS_OPENPYXL else None
CURRENCY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00%'


class ExcelReportExporter:
    """Export financial reports to Excel."""

    def __init__(self, report_type: str, data: Dict[str, Any]):
        self.report_type = report_type
        self.data = data

    def to_excel_response(self, filename: str = "report.xlsx") -> HttpResponse:
        """Generate an Excel file response."""
        if not HAS_OPENPYXL:
            return HttpResponse(
                "openpyxl is not installed.", status=501,
                content_type="text/plain",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = self.report_type.replace("_", " ").title()

        export_method = getattr(
            self, f"_export_{self.report_type.lower()}", None,
        )
        if export_method:
            export_method(ws)
        else:
            self._export_generic(ws)

        self._auto_column_width(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # ── Report-Specific Exports ─────────────────────────────────────

    def _export_trial_balance(self, ws):
        """Export Trial Balance to worksheet."""
        self._add_title(ws, "Trial Balance", 1)
        headers = [
            "Code", "Account Name",
            "Opening Dr", "Opening Cr",
            "Period Dr", "Period Cr",
            "Closing Dr", "Closing Cr",
        ]
        self._add_header_row(ws, headers, 3)

        row = 4
        for group in self.data.get("account_groups", []):
            ws.cell(row=row, column=1, value=group.get("display_name", ""))
            ws.cell(row=row, column=1).font = SUBTOTAL_FONT
            row += 1
            for acct in group.get("accounts", []):
                ws.cell(row=row, column=1, value=acct.get("account_code", ""))
                ws.cell(row=row, column=2, value=acct.get("account_name", ""))
                for i, key in enumerate([
                    "opening_debit", "opening_credit",
                    "period_debit", "period_credit",
                    "closing_debit", "closing_credit",
                ]):
                    cell = ws.cell(row=row, column=3 + i, value=float(acct.get(key, 0)))
                    cell.number_format = CURRENCY_FORMAT
                row += 1

        totals = self.data.get("grand_totals", {})
        self._add_subtotal_row(ws, row, "Grand Total", [
            totals.get("opening_debit", 0), totals.get("opening_credit", 0),
            totals.get("period_debit", 0), totals.get("period_credit", 0),
            totals.get("closing_debit", 0), totals.get("closing_credit", 0),
        ], start_col=3)

    def _export_profit_loss(self, ws):
        """Export P&L to worksheet."""
        self._add_title(ws, "Profit & Loss Statement", 1)
        headers = ["Account", "Amount"]
        self._add_header_row(ws, headers, 3)

        row = 4
        sections = [
            ("Revenue", "revenue"),
            ("Cost of Goods Sold", "cost_of_goods_sold"),
            ("Operating Expenses", "operating_expenses"),
            ("Other Income", "other_income"),
            ("Other Expenses", "other_expenses"),
        ]
        for label, key in sections:
            section = self.data.get(key, {})
            ws.cell(row=row, column=1, value=label).font = SUBTOTAL_FONT
            row += 1
            for acct in section.get("accounts", []):
                ws.cell(row=row, column=1, value=f"  {acct.get('account_code', '')} — {acct.get('account_name', '')}")
                cell = ws.cell(row=row, column=2, value=float(acct.get("balance", 0)))
                cell.number_format = CURRENCY_FORMAT
                row += 1
            ws.cell(row=row, column=1, value=f"Total {label}").font = SUBTOTAL_FONT
            cell = ws.cell(row=row, column=2, value=float(section.get("total", 0)))
            cell.number_format = CURRENCY_FORMAT
            cell.font = SUBTOTAL_FONT
            row += 2

        for label, key, field in [
            ("Gross Profit", "gross_profit", "amount"),
            ("Operating Income", "operating_income", "amount"),
            ("Net Income", "net_income", "amount"),
        ]:
            ws.cell(row=row, column=1, value=label).font = SUBTOTAL_FONT
            cell = ws.cell(row=row, column=2, value=float(self.data.get(key, {}).get(field, 0)))
            cell.number_format = CURRENCY_FORMAT
            cell.font = SUBTOTAL_FONT
            row += 1

    def _export_balance_sheet(self, ws):
        """Export Balance Sheet to worksheet."""
        self._add_title(ws, "Balance Sheet", 1)
        headers = ["Account", "Amount"]
        self._add_header_row(ws, headers, 3)

        row = 4
        assets = self.data.get("assets", {})
        liabilities = self.data.get("liabilities", {})
        equity = self.data.get("equity", {})

        for label, section in [
            ("Current Assets", assets.get("current_assets", {})),
            ("Fixed Assets", assets.get("fixed_assets", {})),
            ("Current Liabilities", liabilities.get("current_liabilities", {})),
            ("Long-Term Liabilities", liabilities.get("long_term_liabilities", {})),
            ("Equity", equity),
        ]:
            ws.cell(row=row, column=1, value=label).font = SUBTOTAL_FONT
            row += 1
            for acct in section.get("accounts", []):
                ws.cell(row=row, column=1, value=f"  {acct.get('account_code', '')} — {acct.get('account_name', '')}")
                cell = ws.cell(row=row, column=2, value=float(acct.get("balance", 0)))
                cell.number_format = CURRENCY_FORMAT
                row += 1
            row += 1

        ws.cell(row=row, column=1, value="TOTAL ASSETS").font = SUBTOTAL_FONT
        cell = ws.cell(row=row, column=2, value=float(assets.get("total_assets", 0)))
        cell.number_format = CURRENCY_FORMAT
        cell.font = SUBTOTAL_FONT
        row += 1
        ws.cell(row=row, column=1, value="TOTAL LIABILITIES & EQUITY").font = SUBTOTAL_FONT
        cell = ws.cell(row=row, column=2, value=float(self.data.get("total_liabilities_equity", 0)))
        cell.number_format = CURRENCY_FORMAT
        cell.font = SUBTOTAL_FONT

    def _export_general_ledger(self, ws):
        """Export General Ledger to worksheet."""
        self._add_title(ws, "General Ledger", 1)
        headers = ["Date", "Entry #", "Description", "Debit", "Credit", "Balance"]
        row = 3

        for acct in self.data.get("accounts", []):
            ws.cell(row=row, column=1, value=f"{acct['account_code']} — {acct['account_name']}").font = SUBTOTAL_FONT
            row += 1
            self._add_header_row(ws, headers, row)
            row += 1
            # Opening balance
            ws.cell(row=row, column=3, value="Opening Balance")
            cell = ws.cell(row=row, column=6, value=float(acct.get("opening_balance", 0)))
            cell.number_format = CURRENCY_FORMAT
            row += 1
            for txn in acct.get("transactions", []):
                ws.cell(row=row, column=1, value=txn.get("date", ""))
                ws.cell(row=row, column=2, value=txn.get("entry_number", ""))
                ws.cell(row=row, column=3, value=txn.get("description", ""))
                for col_i, key in [(4, "debit_amount"), (5, "credit_amount"), (6, "running_balance")]:
                    cell = ws.cell(row=row, column=col_i, value=float(txn.get(key, 0)))
                    cell.number_format = CURRENCY_FORMAT
                row += 1
            row += 1

    def _export_generic(self, ws):
        """Fallback: export raw data keys."""
        ws.cell(row=1, column=1, value="Report Data")
        row = 2
        for key, value in self.data.items():
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=2, value=str(value)[:32000])
            row += 1

    # ── Shared Utilities ────────────────────────────────────────────

    def _add_title(self, ws, title: str, row: int):
        """Add a bold title row."""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = TITLE_FONT

    def _add_header_row(self, ws, headers: List[str], row: int):
        """Add styled header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    def _add_subtotal_row(self, ws, row, label, values, start_col=2):
        """Add a subtotal row with values."""
        ws.cell(row=row, column=1, value=label).font = SUBTOTAL_FONT
        for i, val in enumerate(values):
            cell = ws.cell(row=row, column=start_col + i, value=float(val))
            cell.number_format = CURRENCY_FORMAT
            cell.font = SUBTOTAL_FONT
            cell.border = DOUBLE_BORDER

    def _auto_column_width(self, ws):
        """Auto-fit column widths."""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
