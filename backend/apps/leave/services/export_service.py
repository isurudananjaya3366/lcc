"""Leave Export Service for the Leave Management app.

Generates Excel, PDF, and CSV exports from report data.
"""

import csv
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class LeaveExportService:
    """Service class for exporting leave reports to various formats."""

    def __init__(self, tenant=None):
        self.tenant = tenant

    # ── Core Export Methods ──────────────────────────────────

    def export_to_excel(self, report_data, report_type):
        """Generate Excel workbook from report data.

        Args:
            report_data: Dict from a report service method.
            report_type: String identifying the report type.

        Returns:
            BytesIO containing the Excel file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV.")
            return self.export_to_csv(report_data, report_type)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.replace("_", " ").title()

        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        total_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

        dispatch = {
            "balance_summary": self._excel_balance_summary,
            "leave_history": self._excel_leave_history,
            "department_leave": self._excel_department_report,
            "leave_type_usage": self._excel_leave_type_usage,
            "pending_approvals": self._excel_pending_approvals,
            "expiring_leaves": self._excel_expiring_leaves,
        }

        handler = dispatch.get(report_type)
        if handler:
            handler(ws, report_data, header_fill, header_font, data_font, total_fill)
        else:
            self._excel_generic(ws, report_data, header_fill, header_font, data_font)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_to_csv(self, report_data, report_type):
        """Generate CSV from report data.

        Args:
            report_data: Dict from a report service method.
            report_type: String identifying the report type.

        Returns:
            BytesIO containing the CSV file.
        """
        output = io.BytesIO()
        text_output = io.StringIO()
        writer = csv.writer(text_output)

        dispatch = {
            "balance_summary": self._csv_balance_summary,
            "leave_history": self._csv_leave_history,
            "pending_approvals": self._csv_pending_approvals,
            "expiring_leaves": self._csv_expiring_leaves,
        }

        handler = dispatch.get(report_type)
        if handler:
            handler(writer, report_data)
        else:
            writer.writerow(["Report Type", report_type])
            writer.writerow(["Generated At", report_data.get("generated_at", "")])

        output.write(text_output.getvalue().encode("utf-8-sig"))
        output.seek(0)
        return output

    def export_to_pdf(self, report_data, report_type):
        """Generate PDF from report data.

        Args:
            report_data: Dict from a report service method.
            report_type: String identifying the report type.

        Returns:
            BytesIO containing the PDF file.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            logger.warning("reportlab not installed, falling back to CSV.")
            return self.export_to_csv(report_data, report_type)

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title = report_type.replace("_", " ").title()
        elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
        elements.append(Spacer(1, 12))

        generated = report_data.get("generated_at", "")
        elements.append(
            Paragraph(f"Generated: {generated}", styles["Normal"])
        )
        elements.append(Spacer(1, 12))

        # Build table data based on report type
        table_data = self._build_pdf_table_data(report_data, report_type)
        if table_data:
            table = Table(table_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                    ]
                )
            )
            elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output

    # ── Private Helpers ──────────────────────────────────────

    @staticmethod
    def _generate_filename(report_type, ext):
        """Generate a safe filename with timestamp."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{report_type}_{timestamp}.{ext}"

    # ── Excel Handlers ───────────────────────────────────────

    def _excel_balance_summary(self, ws, data, hfill, hfont, dfont, tfill):
        headers = ["Employee", "Department", "Leave Type", "Allocated", "Used", "Pending", "Available", "Utilization %"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont

        row = 2
        for emp in data.get("employees", []):
            for bal in emp.get("balances", []):
                ws.cell(row=row, column=1, value=emp["name"]).font = dfont
                ws.cell(row=row, column=2, value=emp["department"]).font = dfont
                ws.cell(row=row, column=3, value=bal["leave_type_name"]).font = dfont
                ws.cell(row=row, column=4, value=float(bal["allocated"])).font = dfont
                ws.cell(row=row, column=5, value=float(bal["used"])).font = dfont
                ws.cell(row=row, column=6, value=float(bal["pending"])).font = dfont
                ws.cell(row=row, column=7, value=float(bal["available"])).font = dfont
                ws.cell(row=row, column=8, value=bal["utilization_percentage"]).font = dfont
                row += 1

    def _excel_leave_history(self, ws, data, hfill, hfont, dfont, tfill):
        headers = ["Leave Type", "Start Date", "End Date", "Days", "Status", "Applied", "Reason"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont

        row = 2
        for req in data.get("requests", []):
            ws.cell(row=row, column=1, value=req["leave_type"]["name"]).font = dfont
            ws.cell(row=row, column=2, value=req["dates"]["start_date"]).font = dfont
            ws.cell(row=row, column=3, value=req["dates"]["end_date"]).font = dfont
            ws.cell(row=row, column=4, value=req["dates"]["total_days"]).font = dfont
            ws.cell(row=row, column=5, value=req["status"]["display"]).font = dfont
            ws.cell(row=row, column=6, value=req["application"]["applied_date"]).font = dfont
            ws.cell(row=row, column=7, value=req["application"]["reason"]).font = dfont
            row += 1

    def _excel_department_report(self, ws, data, hfill, hfont, dfont, tfill):
        headers = ["Employee", "Days Taken", "Requests"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont

        row = 2
        for emp in data.get("employee_details", []):
            ws.cell(row=row, column=1, value=emp["name"]).font = dfont
            ws.cell(row=row, column=2, value=emp["total_days_taken"]).font = dfont
            ws.cell(row=row, column=3, value=emp["requests_count"]).font = dfont
            row += 1

    def _excel_leave_type_usage(self, ws, data, hfill, hfont, dfont, tfill):
        headers = ["Metric", "Value"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont

        stats = data.get("usage_statistics", {})
        row = 2
        for key, val in stats.items():
            ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = dfont
            ws.cell(row=row, column=2, value=str(val)).font = dfont
            row += 1

    def _excel_pending_approvals(self, ws, data, hfill, hfont, dfont, tfill):
        headers = ["Employee", "Leave Type", "Start", "End", "Days", "Urgency", "Pending Days"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont

        row = 2
        for req in data.get("pending_requests", []):
            ws.cell(row=row, column=1, value=req["employee"]["name"]).font = dfont
            ws.cell(row=row, column=2, value=req["leave_type"]["name"]).font = dfont
            ws.cell(row=row, column=3, value=req["dates"]["start_date"]).font = dfont
            ws.cell(row=row, column=4, value=req["dates"]["end_date"]).font = dfont
            ws.cell(row=row, column=5, value=req["dates"]["total_days"]).font = dfont
            ws.cell(row=row, column=6, value=req["urgency_level"]).font = dfont
            ws.cell(row=row, column=7, value=req["timeline"]["days_pending"]).font = dfont
            row += 1

    def _excel_expiring_leaves(self, ws, data, hfill, hfont, dfont, tfill):
        headers = ["Employee", "Leave Type", "Available", "Expiry Date", "Days Until", "Urgency"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont

        row = 2
        for item in data.get("expiring_balances", []):
            ws.cell(row=row, column=1, value=item["employee"]["name"]).font = dfont
            ws.cell(row=row, column=2, value=item["leave_type"]["name"]).font = dfont
            ws.cell(row=row, column=3, value=item["balance"]["available_days"]).font = dfont
            ws.cell(row=row, column=4, value=item["expiry"]["expiry_date"]).font = dfont
            ws.cell(row=row, column=5, value=item["expiry"]["days_until_expiry"]).font = dfont
            ws.cell(row=row, column=6, value=item["urgency_level"]).font = dfont
            row += 1

    def _excel_generic(self, ws, data, hfill, hfont, dfont):
        ws.cell(row=1, column=1, value="Key").fill = hfill
        ws.cell(row=1, column=1).font = hfont
        ws.cell(row=1, column=2, value="Value").fill = hfill
        ws.cell(row=1, column=2).font = hfont
        row = 2
        for key, val in data.items():
            if not isinstance(val, (dict, list)):
                ws.cell(row=row, column=1, value=str(key)).font = dfont
                ws.cell(row=row, column=2, value=str(val)).font = dfont
                row += 1

    # ── CSV Handlers ─────────────────────────────────────────

    def _csv_balance_summary(self, writer, data):
        writer.writerow(["Employee", "Department", "Leave Type", "Allocated", "Used", "Pending", "Available", "Utilization %"])
        for emp in data.get("employees", []):
            for bal in emp.get("balances", []):
                writer.writerow([
                    emp["name"], emp["department"], bal["leave_type_name"],
                    bal["allocated"], bal["used"], bal["pending"],
                    bal["available"], bal["utilization_percentage"],
                ])

    def _csv_leave_history(self, writer, data):
        writer.writerow(["Leave Type", "Start Date", "End Date", "Days", "Status", "Applied", "Reason"])
        for req in data.get("requests", []):
            writer.writerow([
                req["leave_type"]["name"], req["dates"]["start_date"],
                req["dates"]["end_date"], req["dates"]["total_days"],
                req["status"]["display"], req["application"]["applied_date"],
                req["application"]["reason"],
            ])

    def _csv_pending_approvals(self, writer, data):
        writer.writerow(["Employee", "Leave Type", "Start", "End", "Days", "Urgency", "Pending Days"])
        for req in data.get("pending_requests", []):
            writer.writerow([
                req["employee"]["name"], req["leave_type"]["name"],
                req["dates"]["start_date"], req["dates"]["end_date"],
                req["dates"]["total_days"], req["urgency_level"],
                req["timeline"]["days_pending"],
            ])

    def _csv_expiring_leaves(self, writer, data):
        writer.writerow(["Employee", "Leave Type", "Available", "Expiry Date", "Days Until", "Urgency"])
        for item in data.get("expiring_balances", []):
            writer.writerow([
                item["employee"]["name"], item["leave_type"]["name"],
                item["balance"]["available_days"], item["expiry"]["expiry_date"],
                item["expiry"]["days_until_expiry"], item["urgency_level"],
            ])

    # ── PDF Helpers ──────────────────────────────────────────

    def _build_pdf_table_data(self, report_data, report_type):
        """Build a list-of-lists for PDF table rendering."""
        if report_type == "balance_summary":
            rows = [["Employee", "Department", "Leave Type", "Allocated", "Used", "Available"]]
            for emp in report_data.get("employees", []):
                for bal in emp.get("balances", []):
                    rows.append([
                        emp["name"], emp["department"], bal["leave_type_name"],
                        str(bal["allocated"]), str(bal["used"]), str(bal["available"]),
                    ])
            return rows

        if report_type == "pending_approvals":
            rows = [["Employee", "Leave Type", "Start", "End", "Urgency"]]
            for req in report_data.get("pending_requests", []):
                rows.append([
                    req["employee"]["name"], req["leave_type"]["name"],
                    req["dates"]["start_date"], req["dates"]["end_date"],
                    req["urgency_level"],
                ])
            return rows

        return None
